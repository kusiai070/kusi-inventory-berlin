import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ReportGenerator:
    """Clase utilitaria para generar reportes en PDF y Excel"""

    @staticmethod
    def generate_excel(data: List[Dict[str, Any]], title: str, columns: List[Dict[str, str]]) -> io.BytesIO:
        """
        Genera un archivo Excel en memoria.
        
        Args:
            data: Lista de diccionarios con los datos.
            title: Título del reporte.
            columns: Lista de diccionarios definiendo columnas [{'key': 'key_in_data', 'header': 'Column Header'}]
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        centered_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(size=14, bold=True)
        cell.alignment = centered_alignment
        
        # Fecha generación
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        # Headers
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=4, column=col_idx, value=col_def['header'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = thin_border
            
            # Ajustar ancho (aproximado)
            ws.column_dimensions[chr(64 + col_idx)].width = 20
        
        # Datos
        row_idx = 5
        for item in data:
            for col_idx, col_def in enumerate(columns, 1):
                key = col_def['key']
                value = item.get(key, '')
                
                # Formatear números
                if isinstance(value, float):
                    value = round(value, 2)
                    
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Alineación
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
            row_idx += 1
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_pdf(data: List[Dict[str, Any]], title: str, columns: List[Dict[str, str]], summary_info: Optional[Dict] = None) -> io.BytesIO:
        """
        Genera un archivo PDF en memoria.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=10,
            alignment=1 # Center
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Resumen (si existe)
        if summary_info:
            elements.append(Paragraph("Resumen:", styles['Heading3']))
            for key, value in summary_info.items():
                elements.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))
            elements.append(Spacer(1, 15))
        
        # Tabla de Datos
        # Headers
        table_data = [[col['header'] for col in columns]]
        
        # Rows
        for item in data:
            row = []
            for col in columns:
                val = item.get(col['key'], '')
                if isinstance(val, float):
                    val = f"{val:.2f}"
                row.append(str(val))
            table_data.append(row)
            
        # Crear tabla
        col_widths = [None] * len(columns) # Auto width
        t = Table(table_data, repeatRows=1)
        
        # Estilo Tabla
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'), # Primera columna a la izquierda
        ]))
        
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer
